# -*- coding: utf-8 -*-
"""哲學 대체어 치환 재현 — 우리 5문장 묶음 짝에서.

직접 1915↔1924 묶음 짝(≥0.1) 중 *1915 묶음에 哲學*이 든 짝을 뽑아,
대응 1924 묶음의 哲學 유무(소거 여부)와 그 자리 대체 어휘를 raw로 대조.
소스: BK_IT_1915 · BK_YD_1924 xlsx (kr_text+n_chunk_id) + norm tokens(짝 산출).
"""
import sys, csv
from collections import defaultdict
from pathlib import Path
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

OUT = Path(r"C:\hp_data\0_tnt\hyeonsang\_ocr_experiments\unified_db_2026-05-19\output")
NORM = OUT / "norm"
THR = 0.10

def load_tok(fn):
    sets = defaultdict(set)
    with open(NORM / fn, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            if r["n_chunk_id"] and r["token"]:
                sets[r["n_chunk_id"]].add(r["token"])
    return sets

def load_text(fn):
    """n_chunk_id -> [행 텍스트들] (kr_text)."""
    rows = defaultdict(list)
    wb = openpyxl.load_workbook(OUT / fn, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it)); ti = hdr.index("kr_text"); ni = hdr.index("n_chunk_id")
    for r in it:
        if r[ni] and r[ti]:
            rows[str(r[ni])].append(str(r[ti]))
    wb.close()
    return rows

s15, s24 = load_tok("tokens_1915.csv"), load_tok("tokens_1924.csv")
t15 = load_text("BK_IT_1915_PR_v1.4.xlsx")
t24 = load_text("BK_YD_1924_IY_v1.3.xlsx")

inv = defaultdict(set)
for cid, s in s15.items():
    for t in s:
        inv[t].add(cid)

# 직접 짝(≥0.1)
pairs = []
for c24, A in s24.items():
    cand = set()
    for t in A:
        cand |= inv.get(t, set())
    for c15 in cand:
        B = s15[c15]
        j = len(A & B) / len(A | B)
        if j >= THR:
            pairs.append((j, c24, c15))
pairs.sort(reverse=True)

def has_phil(rowlist):
    return any("哲學" in r or "哲学" in r for r in rowlist)
def phil_rows(rowlist):
    return [r for r in rowlist if "哲學" in r or "哲学" in r]

# 1915 묶음에 哲學 든 짝
hits = [(j, c24, c15) for j, c24, c15 in pairs if has_phil(t15.get(c15, []))]
soz = [h for h in hits if not has_phil(t24.get(h[1], []))]
print(f"{'='*66}\n哲學 대체어 치환 — 직접 묶음 짝(≥{THR}) 중 1915측 哲學 보유\n{'='*66}")
print(f"1915측 哲學 보유 짝 {len(hits)} · 그 중 1924측 哲學 *소거* {len(soz)} "
      f"· 哲學 유지 {len(hits)-len(soz)}")

print(f"\n--- 소거 사례 (1915 哲學 → 1924 哲學 없음) ---")
for j, c24, c15 in soz[:6]:
    pr = phil_rows(t15[c15])
    snip15 = pr[0][:74] if pr else ""
    snip24 = " ".join(t24[c24])[:120]
    print(f"\n[j={j:.3f}]  1915 {c15}  →  1924 {c24}")
    print(f"  1915(哲學): …{snip15}…")
    print(f"  1924(대응): {snip24}")

print(f"\n--- 哲學 유지 사례 (대조군) ---")
for j, c24, c15 in [h for h in hits if has_phil(t24.get(h[1], []))][:3]:
    pr15 = phil_rows(t15[c15]); pr24 = phil_rows(t24[c24])
    print(f"\n[j={j:.3f}]  1915 {c15}  →  1924 {c24}")
    print(f"  1915(哲學): …{(pr15[0] if pr15 else '')[:70]}…")
    print(f"  1924(哲學): …{(pr24[0] if pr24 else '')[:70]}…")

# -*- coding: utf-8 -*-
"""ⓐ 3구분법 「종교·철학·과학」 존재-타임라인 — raw 텍스트 tight 공기 추적.

소스: MA_YD_10-20_GB/WB.xlsx raw_text. 기사별 텍스트를 이어 붙여 종교·철학·과학
세 개념(한자+한글)이 최소 몇 자 안에 모이는지(min-span) 계산. tight=≤12자(리스트형
'宗敎哲學科學'), loose=≤40자(분산 사용). 기사 단위 *존재* 추적(소수라 빈도율 아님).
출력: 표 + appendix/3구분법_타임라인.{png,svg}
"""
import sys, csv
from pathlib import Path
from collections import defaultdict
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

OUTDIR = Path(r"C:\hp_data\0_tnt\hyeonsang\_ocr_experiments\unified_db_2026-05-19\output")
META = Path(__file__).parent / "mag_meta.csv"
APP = Path(__file__).parent.parent / "appendix"
JONG = ("宗敎", "종교"); CHOL = ("哲學", "철학"); GWA = ("科學", "과학")
TIGHT, LOOSE = 12, 40

meta = {}
with open(META, encoding="utf-8-sig") as f:
    for r in csv.DictReader(f):
        meta[(r["src"], r["c"])] = r["date"][:4]

def starts(text, forms):
    out = []
    for fm in forms:
        i = text.find(fm)
        while i != -1:
            out.append(i)
            i = text.find(fm, i + 1)
    return out

def min_span(text):
    """종교·철학·과학 각 1개 이상 포함하는 최소 char 윈도우 폭. 없으면 None."""
    evs = ([(p, 0) for p in starts(text, JONG)] +
           [(p, 1) for p in starts(text, CHOL)] +
           [(p, 2) for p in starts(text, GWA)])
    if not all(any(t == k for _, t in evs) for k in (0, 1, 2)):
        return None
    evs.sort()
    best = 10**9
    cnt = [0, 0, 0]; have = 0; lo = 0
    for hi in range(len(evs)):
        if cnt[evs[hi][1]] == 0:
            have += 1
        cnt[evs[hi][1]] += 1
        while have == 3:
            best = min(best, evs[hi][0] - evs[lo][0])
            cnt[evs[lo][1]] -= 1
            if cnt[evs[lo][1]] == 0:
                have -= 1
            lo += 1
    return best

# 기사별 텍스트 결합 → min-span
art_text = defaultdict(list)
for fn, ven in (("MA_YD_10-20_GB.xlsx", "개벽"), ("MA_YD_10-20_WB.xlsx", "월보")):
    wb = openpyxl.load_workbook(OUTDIR / fn, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    hdr = list(next(rows)); li = hdr.index("local_id"); ti = hdr.index("raw_text")
    for r in rows:
        if r[ti]:
            art_text[(ven, str(r[li]).split("-")[0])].append(str(r[ti]))
    wb.close()

hits = []   # (ven, c, year, span)
for (ven, c), lines in art_text.items():
    sp = min_span(" ".join(lines))
    if sp is not None and sp <= LOOSE:
        hits.append((ven, c, meta.get((ven, c), "?"), sp))
hits.sort(key=lambda x: (x[0], x[2], x[3]))

print(f"{'='*64}\nⓐ「종교철학과학」 공기 기사 (span≤{LOOSE}자)\n{'='*64}")
print(f"{'매체':4}{'기사':5}{'연도':6}{'min-span':>9}  분류")
for ven, c, yr, sp in hits:
    print(f"{ven:4}{c:5}{yr:6}{sp:>9}  {'TIGHT' if sp <= TIGHT else 'loose'}")

# 연도×매체 tight 기사수
tcount = defaultdict(int); lcount = defaultdict(int)
for ven, c, yr, sp in hits:
    if yr == "?":
        continue
    if sp <= TIGHT:
        tcount[(ven, int(yr))] += 1
    else:
        lcount[(ven, int(yr))] += 1
print(f"\n연도별 TIGHT 기사수: " +
      ", ".join(f"{v}{y}={n}" for (v, y), n in sorted(tcount.items())))

# 이벤트 타임라인
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
plt.rcParams["font.family"] = "Malgun Gothic"
plt.rcParams["axes.unicode_minus"] = False
LANE = {"월보": 1.0, "개벽": 0.0}
fig, ax = plt.subplots(figsize=(10.2, 4.3))
ax.axvspan(1915.5, 1918.5, color="#f0f0f2", zorder=0)
ax.axvspan(1918.5, 1921.5, color="#fcf3d9", zorder=0)
ax.text(1917.0, 1.62, "잠복", ha="center", fontsize=10, color="#888", style="italic")
ax.text(1920.0, 1.62, "재대두", ha="center", fontsize=10, color="#b8860b", fontweight="bold")
ax.text(1913.0, 1.62, "등장", ha="center", fontsize=10, color="#777")
ax.text(1924.0, 1.62, "리스트형 소멸 (분산만)", ha="center", fontsize=9, color="#999", style="italic")
# (lane,year)별 지터
bucket = defaultdict(list)
for ven, c, yr, sp in hits:
    if yr != "?":
        bucket[(ven, int(yr))].append((c, sp))
for (ven, y), items in bucket.items():
    n = len(items)
    for k, (c, sp) in enumerate(sorted(items, key=lambda x: x[1])):
        dx = (k - (n - 1) / 2) * 0.16
        tight = sp <= TIGHT
        ax.scatter(y + dx, LANE[ven], s=150 if tight else 70,
                   facecolors="#c0392b" if tight else "none",
                   edgecolors="#c0392b" if tight else "#999",
                   linewidths=1.4, zorder=4)
        ax.annotate(c, (y + dx, LANE[ven]), xytext=(0, 11 if tight else -15),
                    textcoords="offset points", ha="center",
                    fontsize=7.6, color="#333" if tight else "#999",
                    fontweight="bold" if tight else "normal")
ax.set_yticks([0, 1]); ax.set_yticklabels(["개벽", "월보"], fontsize=11)
ax.set_ylim(-0.6, 1.85); ax.set_xlim(1910.5, 1926.5)
ax.set_xticks(range(1911, 1927))
ax.set_xlabel("연도", fontsize=10.5)
for s in ("top", "right", "left"):
    ax.spines[s].set_visible(False)
ax.tick_params(left=False)
from matplotlib.lines import Line2D
ax.legend(handles=[
    Line2D([0], [0], marker="o", color="w", markerfacecolor="#c0392b",
           markersize=11, label="TIGHT (리스트형 ≤12자)"),
    Line2D([0], [0], marker="o", color="w", markerfacecolor="none",
           markeredgecolor="#999", markersize=9, label="loose (분산 ≤40자)")],
    loc="lower right", fontsize=8.8, frameon=False)
ax.set_title("ⓐ 3구분법 「종교·철학·과학」 출현 — 1915 등장 → 잠복 → 1919–21 재대두 → 소멸\n"
             "(허수 2011 line 641 잠복·재대두 서사의 raw 텍스트 실측)",
             fontsize=12, fontweight="bold", loc="left")
fig.tight_layout()
fig.savefig(APP / "3구분법_타임라인.svg", bbox_inches="tight")
fig.savefig(APP / "3구분법_타임라인.png", dpi=185, bbox_inches="tight")
print(f"saved {APP / '3구분법_타임라인.png'} / .svg")

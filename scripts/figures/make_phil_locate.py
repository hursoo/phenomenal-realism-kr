# -*- coding: utf-8 -*-
"""哲學 6회 위치 추적 + 역상관 완성 — 1924 책 전체 항별 [참조밀도 ↔ 哲學 빈도].

(1) 직접 짝(≥0.1) 항별 참조 수. (2) tokens_1924 항별 哲學 수.
(3) BK_YD_1924 xlsx에서 哲學/철학 raw_text·위치 6건 추출(맥락 확인).
"""
import sys, csv, re
from collections import defaultdict
from pathlib import Path
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

NORM = Path(r"C:\hp_data\0_tnt\hyeonsang\_ocr_experiments\unified_db_2026-05-19\output\norm")
XLSX = Path(r"C:\hp_data\0_tnt\hyeonsang\_ocr_experiments\unified_db_2026-05-19\output\BK_YD_1924_IY_v1.3.xlsx")
THR = 0.10
PHIL = {"哲學", "철학"}

def parse(cid):
    m = re.match(r"(C\d+)-(S\d+)-(I\d+)", cid or "")
    return m.groups() if m else None

def load(fn):
    sets = defaultdict(set)
    with open(NORM / fn, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            if r["n_chunk_id"] and r["token"]:
                sets[r["n_chunk_id"]].add(r["token"])
    return sets

s15, s24 = load("tokens_1915.csv"), load("tokens_1924.csv")
inv = defaultdict(set)
for cid, s in s15.items():
    for t in s:
        inv[t].add(cid)

# 항(C,S,I)별 참조 수
ref_si = defaultdict(int)
for c24, A in s24.items():
    si = parse(c24)
    if not si:
        continue
    cand = set()
    for t in A:
        cand |= inv.get(t, set())
    for c15 in cand:
        if len(A & s15[c15]) / len(A | s15[c15]) >= THR:
            ref_si[si] += 1

# 항별 哲學 수
phil_si = defaultdict(int)
with open(NORM / "tokens_1924.csv", encoding="utf-8-sig") as f:
    for r in csv.DictReader(f):
        if r["n_chunk_id"] and r["token"] in PHIL:
            si = parse(r["n_chunk_id"])
            if si:
                phil_si[si] += 1

print(f"{'='*60}\n역상관 — 1924 항별 [참조밀도 ↔ 哲學] (참조 많은 순)\n{'='*60}")
allkeys = sorted(set(ref_si) | set(phil_si), key=lambda k: -ref_si.get(k, 0))
print(f"{'항(C-S-I)':16}{'참조':>5}{'哲學':>5}")
for k in allkeys:
    if ref_si.get(k, 0) >= 3 or phil_si.get(k, 0) > 0:
        print(f"{'-'.join(k):16}{ref_si.get(k,0):>5}{phil_si.get(k,0):>5}")

# 哲學 보유 항 vs 최다참조 항 대비
phil_keys = [k for k in phil_si if phil_si[k] > 0]
top_ref = sorted(ref_si.items(), key=lambda x: -x[1])[:8]
print(f"\n哲學 보유 항 {len(phil_keys)}곳의 참조 합 = {sum(ref_si.get(k,0) for k in phil_keys)} "
      f"(哲學 총 {sum(phil_si.values())}회)")
print(f"최다참조 상위 8항의 哲學 합 = {sum(phil_si.get(k,0) for k,_ in top_ref)} "
      f"(참조 총 {sum(n for _,n in top_ref)})")

# 哲學 raw_text 6건 (xlsx)
print(f"\n{'='*60}\n哲學/철학 raw_text 출현 (BK_YD_1924) + 위치·그 자리 참조 수\n{'='*60}")
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = ws.iter_rows(values_only=True)
hdr = list(next(rows)); ti = hdr.index("kr_text"); ni = hdr.index("n_chunk_id")
n = 0
for r in rows:
    txt = r[ti]
    if txt and ("哲學" in str(txt) or "철학" in str(txt)):
        n += 1
        si = parse(str(r[ni]))
        ref = ref_si.get(si, 0) if si else "?"
        loc = '-'.join(si) if si else str(r[ni])
        print(f"  [{loc} · 참조 {ref}] {str(txt)[:64]}")
wb.close()
print(f"\n→ raw_text 哲學 출현 {n}행")
